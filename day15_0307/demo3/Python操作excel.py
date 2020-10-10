import openpyxl

# 1、将指定的excel  ,加载为一个workbook对象
wb = openpyxl.load_workbook("cases.xlsx")

# 2、选中工作簿中的表单对象
biao_dan = wb["大胖"]

# 3、读取指定行列的格子对象
a1 = biao_dan.cell(row =1,column=2)

# 4、 读取格子中的数据
res_a1 = a1.value
# print(res_a1)    # 222


# 5、写入数据，在对应的表单中写入。  注意：写入的时候需要关闭掉excel
biao_dan.cell(row =4,column=1,value="这位突突突突")
# 将工作簿对象保存为文件
wb.save("cases.xlsx")


# 6、获取表单中的最大行
max_row = biao_dan.max_row
# print(max_row)
# 7、获取表单中的最大列
max_co = biao_dan.max_column
# print(max_co)

# 8、rows  按行获取表单中的所有的格子对象，每行的内容，放在一个元祖中。
res = list(biao_dan.rows)
# print(res)

for item in res:
    # print(item)
    for j in item:
        print(j.value)

# 9、columns 按列获取表单中的所有格子的对象，每列的内容太放在一个元祖中。
res2 = list(biao_dan.columns)
print(res2)

"""
这是打印出来的res:
[(<Cell '大胖'.A1>, <Cell '大胖'.B1>, <Cell '大胖'.C1>, <Cell '大胖'.D1>), 
(<Cell '大胖'.A2>, <Cell '大胖'.B2>, <Cell '大胖'.C2>, <Cell '大胖'.D2>), 
(<Cell '大胖'.A3>, <Cell '大胖'.B3>, <Cell '大胖'.C3>, <Cell '大胖'.D3>), 
(<Cell '大胖'.A4>, <Cell '大胖'.B4>, <Cell '大胖'.C4>, <Cell '大胖'.D4>)]

"""

"""
这是打印出来的item:
(<Cell '大胖'.A1>, <Cell '大胖'.B1>, <Cell '大胖'.C1>, <Cell '大胖'.D1>)
(<Cell '大胖'.A2>, <Cell '大胖'.B2>, <Cell '大胖'.C2>, <Cell '大胖'.D2>)
(<Cell '大胖'.A3>, <Cell '大胖'.B3>, <Cell '大胖'.C3>, <Cell '大胖'.D3>)
(<Cell '大胖'.A4>, <Cell '大胖'.B4>, <Cell '大胖'.C4>, <Cell '大胖'.D4>)

"""
"""
这是打印出来的j :
<Cell '大胖'.A1>
<Cell '大胖'.B1>
<Cell '大胖'.C1>
<Cell '大胖'.D1>
<Cell '大胖'.A2>
<Cell '大胖'.B2>
<Cell '大胖'.C2>
<Cell '大胖'.D2>
<Cell '大胖'.A3>
<Cell '大胖'.B3>
<Cell '大胖'.C3>
<Cell '大胖'.D3>
<Cell '大胖'.A4>
<Cell '大胖'.B4>
<Cell '大胖'.C4>
<Cell '大胖'.D4>
"""

